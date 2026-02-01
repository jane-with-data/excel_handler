"""Excel sheet formatting module.

This module provides functionality to format Excel worksheets with styles,
column widths, filters, and freeze panes based on configuration.

Classes:
    SetSheetFormatter: Main class for applying formatting to workbook sheets.

Example:
    >>> from openpyxl import Workbook
    >>> from services.excel_handler.set_sheet_formatter import SetSheetFormatter
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> formatter = SetSheetFormatter(wb)
    >>> formatter.run_pipeline(ws, df)  # Apply all formatting
"""
from typing import Optional, List
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from services.excel_handler.style_builder import build_bulk_named_style
from shared.constants import (
    MAP_DATA_TYPE_STYLE_NAME_CONFIG,
    DECLARE_DATA_TYPE_CONFIG,
    VISUAL_STYLE_NAME_CONFIG,
    DATA_TYPE_STYLE_NAME_CONFIG,
)
from services.logger.logger import get_logger

logger = get_logger()
# logger = logging.getLogger(__name__)

class SetSheetFormatter:
    """Format Excel worksheets with styles, widths, filters, and freeze panes.
    
    This class applies comprehensive formatting to Excel worksheets including:
    - Named styles (header, body, data types)
    - Column width auto-adjustment
    - Auto-filter configuration
    - Freeze panes setup
    
    Attributes:
        wb (Workbook): The Excel workbook to format.
    
    Example:
        >>> wb = Workbook()
        >>> ws = wb.active
        >>> formatter = SetSheetFormatter(wb)
        >>> formatter.run_pipeline(ws, df)
    """

    def __init__(self, wb: Workbook) -> None:
        """Initialize formatter and register global styles.
        
        Args:
            wb (Workbook): Excel workbook to format.
        
        Raises:
            TypeError: If wb is not a Workbook instance.
        """
        if not isinstance(wb, Workbook):
            raise TypeError(f"Expected Workbook, got {type(wb).__name__}")
        
        self.wb = wb
        self._init_global_styles()
        logger.info("NamedStyle, SetSheetFormatter initialized successfully")
    
    def _init_global_styles(self) -> None:
        """Register all styles into the workbook once.
        
        This method registers both visual format styles and data type format
        styles into the workbook to avoid duplicate style definitions.
        
        Raises:
            Exception: If style registration fails.
        """
        # get list of style_names, data_type_style_names
        try:
            named_style_lst = build_bulk_named_style()
            for item in named_style_lst:
                self.wb.add_named_style(item)
            logger.debug(
                f"Registered {len(named_style_lst)} named_style_lst"
            )
            
        except Exception as e:
            logger.error(f"Error initializing styles: {item} // {e}", exc_info=True)
            raise
    
    def format_style_base_on_data_type(
        self, 
        ws: Worksheet, 
        column_name_lst: List[str], 
        num_rows: int
    ) -> None:
        """Apply data type specific styles to worksheet columns.
        
        Maps column names to data type configurations, then applies
        corresponding styles to each column (excluding header).
        
        Args:
            ws (Worksheet): Worksheet to format.
            column_name_lst (List[str]): List of column names in order.
            num_rows (int): Number of data rows (excluding header).
        
        Raises:
            ValueError: If column_name_lst is empty or num_rows is negative.
        """
        if not column_name_lst:
            logger.warning("Empty column list provided")
            return
        
        if num_rows < 0:
            raise ValueError(f"num_rows must be non-negative, got {num_rows}")
        
        logger.debug(f"Formatting {len(column_name_lst)} columns based on data types")
        
        for col in column_name_lst:
            # Get data type configuration for this column `col`
            cfg_data_type: str = DECLARE_DATA_TYPE_CONFIG.get(col, 'default_data_type')
            cfg_style_name: str = MAP_DATA_TYPE_STYLE_NAME_CONFIG.get(
                cfg_data_type, 
                'default_style'
            )
            
            # Calculate cell range (start from row 2 to skip header)
            col_applied_style: int = column_name_lst.index(col) + 1
            logger.debug(
                f"Column: {col}, Type: {cfg_data_type}, "
                f"Style: {cfg_style_name}, Col Index: {col_applied_style}"
            )
            
            # Apply style to data rows
            self._apply_styles(
                ws, 
                row_idx=2, 
                col_idx=col_applied_style, 
                num_rows=num_rows, 
                num_cols=1, 
                applied_style_name=cfg_style_name
            )
        
        logger.debug("Data type styling completed")

    def format_style_header_body(
        self, 
        ws: Worksheet, 
        num_rows: int, 
        num_cols: int
    ) -> None:
        """Apply header and body styles to worksheet.
        
        Applies header style to first row and body style to data rows.
        
        Args:
            ws (Worksheet): Worksheet to format.
            num_rows (int): Total number of rows including header.
            num_cols (int): Number of columns.
        
        Raises:
            ValueError: If num_rows or num_cols is non-positive.
        """
        if num_rows < 1 or num_cols < 1:
            raise ValueError(
                f"num_rows ({num_rows}) and num_cols ({num_cols}) must be positive"
            )
        
        logger.debug(f"Formatting header and body: {num_rows} rows, {num_cols} cols")
        
        # Format header row (row 1)
        self._apply_styles(
            ws, 
            row_idx=1, 
            col_idx=1, 
            num_rows=1, 
            num_cols=num_cols, 
            applied_style_name="header_style"
        )
        
        # Format body rows (rows 2 to num_rows)
        self._apply_styles(
            ws, 
            row_idx=2, 
            col_idx=1, 
            num_rows=num_rows - 1, 
            num_cols=num_cols, 
            applied_style_name="body_style"
        )
        
        logger.debug("Header and body styling completed")
    
    def _apply_styles(
        self,
        ws: Worksheet,
        row_idx: int,
        col_idx: int,
        num_rows: int,
        num_cols: int,
        applied_style_name: str
    ) -> None:
        """Apply style to a range of cells.
        
        Args:
            ws (Worksheet): Worksheet to modify.
            row_idx (int): Starting row index (1-based).
            col_idx (int): Starting column index (1-based).
            num_rows (int): Number of rows to format.
            num_cols (int): Number of columns to format.
            applied_style_name (str): Style name to apply.
        
        Raises:
            ValueError: If indices are invalid.
        """
        if row_idx < 1 or col_idx < 1:
            raise ValueError(
                f"Row and column indices must be >= 1, "
                f"got row={row_idx}, col={col_idx}"
            )
        
        logger.debug(
            f"Applying '{applied_style_name}' to range: "
            f"rows {row_idx}-{row_idx + num_rows - 1}, "
            f"cols {col_idx}-{col_idx + num_cols - 1}"
        )
        
        for row in range(row_idx, row_idx + num_rows):
            for col in range(col_idx, col_idx + num_cols):
                ws.cell(row=row, column=col).style = applied_style_name

    def auto_adjust_cols(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content.
        
        Calculates the maximum content width for each column and sets
        the column width accordingly, capped at a maximum value.
        
        Args:
            ws (Worksheet): Worksheet to adjust.
        
        Example:
            >>> formatter.auto_adjust_cols(ws)
        """
        try:
            auto_adjust_config =VISUAL_STYLE_NAME_CONFIG.get(
                'auto_adjust_style', 
                {}
            )
            if not auto_adjust_config.get('auto_adjust_width', False):
                logger.debug("Auto-adjust columns disabled in config")
                return
            
            max_width = auto_adjust_config.get('max_column_width', 60)
            logger.debug(f"Auto-adjusting columns (max width: {max_width})")
            
            for col in ws.columns:
                max_len: int = max(
                    [len(str(cell.value or "")) for cell in col],
                    default=5
                )
                adjusted_width: int = min(max_len + 8, max_width)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
            
            logger.debug("Column widths adjusted successfully")
            
        except Exception as e:
            logger.error(f"Error adjusting column widths: {e}", exc_info=True)
            
    def set_filter_mode(self, ws: Worksheet) -> None:
        """Enable or disable auto-filter on worksheet.
        
        Applies auto-filter to all data based on configuration.
        
        Args:
            ws (Worksheet): Worksheet to configure.
        
        Example:
            >>> formatter.set_filter_mode(ws)
        """
        try:
            filter_config = VISUAL_STYLE_NAME_CONFIG.get('filter_style', {})
            if filter_config.get('mode_on', False):
                ws.auto_filter.ref = ws.dimensions
                logger.debug(f"Auto-filter enabled on range: {ws.dimensions}")
            else:
                logger.debug("Auto-filter disabled in config")
        except Exception as e:
            logger.error(f"Error setting filter mode: {e}", exc_info=True)
            
    def set_freeze_panes_style(self, ws: Worksheet) -> None:
        """Enable or disable freeze panes on worksheet.
        
        Freezes rows/columns at specified cell based on configuration.
        
        Args:
            ws (Worksheet): Worksheet to configure.
        
        Example:
            >>> formatter.set_freeze_panes_style(ws)  # Freezes at B2 by default
        """
        try:
            freeze_config = VISUAL_STYLE_NAME_CONFIG.get('freeze_panes_style', {})
            if freeze_config.get('mode_on', False):
                freeze_cell = freeze_config.get('freeze_cell', 'B2')
                ws.freeze_panes = freeze_cell
                logger.debug(f"Freeze panes enabled at: {freeze_cell}")
            else:
                logger.debug("Freeze panes disabled in config")
        except Exception as e:
            logger.error(f"Error setting freeze panes: {e}", exc_info=True)
    
    def run_pipeline(self, ws: Worksheet, df) -> None:
        """Execute full formatting pipeline on worksheet.
        
        Applies all formatting in sequence:
        1. Header and body styles
        2. Data type-specific styles
        3. Auto-filter
        4. Column width adjustment
        5. Freeze panes
        
        Args:
            ws (Worksheet): Worksheet to format.
            df: DataFrame with data (for column and row count).
        
        Raises:
            AttributeError: If df doesn't have required attributes.
        
        Example:
            >>> formatter.run_pipeline(ws, df)
        """
        try:
            logger.debug("Starting formatting pipeline")
            
            # Extract DataFrame dimensions
            col_names: List[str] = list(df.columns)
            num_rows: int = len(df)
            num_cols: int = len(df.columns)
            
            logger.debug(
                f"DataFrame dimensions: {num_rows} rows, {num_cols} columns"
            )
            
            # Apply formatting in sequence
            self.format_style_header_body(ws, num_rows + 1, num_cols)  # +1 for header
            self.format_style_base_on_data_type(ws, col_names, num_rows)
            self.set_filter_mode(ws)
            self.auto_adjust_cols(ws)
            self.set_freeze_panes_style(ws)
            
            logger.debug("Formatting pipeline completed successfully")
            
        except Exception as e:
            logger.error(
                f"Error in formatting pipeline: {e}",
                exc_info=True
            )
            raise